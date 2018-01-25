from optparse import OptionParser
import logging
import openpyxl
from openpyxl import Workbook
import requests, pprint



class SpreadsheetCreator:


    def __init__(self):
        self.logger = logging.getLogger(__name__)

    def generateSpreadsheet(self, baseUri, schemas, dependencies, output):
        values = {}
        try:
            # for each schema, gather the values for the relevant tab(s)
            for schema in schemas:
                v = self._gatherValues(baseUri+schema, dependencies)
                values.update(v)
            # Build the spreadsheet from the retrieved values
            self._buildSpreadsheet(values, output)
        except ValueError as e:
            self.logger.error("Error:" + str(e))
            raise e

    def _gatherValues(self, schema, dependencies):
        # get the schema of HTTP
        req = requests.get(schema)

        # if the schema is successfully retrieved, process it, else return an error message
        if (req.status_code == requests.codes.ok):
            jsonRaw = req.json()

            entities = {}
            entity_title = jsonRaw["title"]

            properties = jsonRaw["properties"]

            values = []

            for prop in properties:
                # if a property has a user_friendly tag, include it as a direct field. This includes ontology module references as these should not be
                # exposed to users
                if ("user_friendly" in properties[prop]):
                    if ("description" in properties[prop]):
                      values.append({"header":properties[prop]["user_friendly"], "description":properties[prop]["description"]})
                # if a property does not include a user_friendly tag but includes a reference, fetch the contents of that reference and add them
                # directly to the properties for this sheet
                elif("$ref" in properties[prop]):
                    module = properties[prop]["$ref"]
                    if "_core" in module or module in dependencies:
                        module_values = self._gatherValues(module, None)
                        for key in module_values.keys():
                            values.extend(module_values[key])
                # if a property has an array of references (potential 1-to-many relationship), gather the properties for the references and format them to become
                # their own spreadsheet tab
                elif("items" in properties[prop] and "$ref" in properties[prop]["items"]):
                    module = properties[prop]["items"]["$ref"]
                    if module in dependencies:
                        module_values = self._gatherValues(module, None)
                        entities.update(module_values)

            entities[entity_title] = values
            return entities

        else:
            self.logger.error(schema + " does not exist")


    def _buildSpreadsheet(self, values, outputLocation):
        wb = Workbook()

        # for each tab entry in the values dictionary, create a new worksheet
        for title in values.keys():

            headers = values[title]

            ws = wb.create_sheet(title=title)
            col = 1

            # put each description in row 1 and each header in row 2, then increment the column index
            for header in headers:
                ws.cell(column=col, row=1, value=header["description"])
                ws.cell(column=col, row=2, value=header["header"])
                col += 1

        # remove the blank worksheet that is automatically created with the spreadsheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb.get_sheet_by_name("Sheet"))

        wb.save(filename=outputLocation)


if __name__ == '__main__':
    parser = OptionParser()
    parser.add_option("-s", "--schema", dest="schema_uri",
                      help="Base schema URI for the metadata")
    parser.add_option("-o", "--output", dest="output",
                      help="Output directory and file where to save the template spreadsheet", metavar="FILE")
    parser.add_option("-t", "--types", dest="schema_types",
                      help="Schema types to include in the spreadsheet")
    parser.add_option("-i", "--include", dest="include",
                      help="Schema modules to include in the spreadsheet")


    (options, args) = parser.parse_args()

    # if not options.schema_uri:
    #     print ("You must supply a base schema URI for the metadata")
    #     exit(2)

    schema_types = options.schema_types.split(",")
    dependencies = options.include.split(",")

    for index, dependency in enumerate(dependencies):
        dependencies[index] = options.schema_uri+dependency

    generator = SpreadsheetCreator()
    generator.generateSpreadsheet(options.schema_uri, schema_types, dependencies, options.output)



# Example run:
# -s "https://raw.githubusercontent.com/HumanCellAtlas/metadata-schema/v5_prototype/json_schema/"
# -t ["type/biomaterial/organism.json", "type/process/sequencing/library_preparation_process.json"]
# -i ["module/biomaterial/homo_sapiens_specific.json", "module/biomateral/familial_relationship.json", "module/process/sequencing/barcode.json"]
# -o "/Users/dwelter/Development/HCA/metadata-schema/src/spreadsheet_test.xlsx"
#